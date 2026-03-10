import streamlit as st
import openpyxl
import io
import pandas as pd

st.set_page_config(
    page_title="在庫管理データ 抽出アプリ",
    page_icon="📦",
    layout="wide"
)

# ── スタイル ──────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main { padding-top: 1rem; }
    .stDataFrame { font-size: 13px; }
    div[data-testid="metric-container"] {
        background: #f0f4f9;
        border-radius: 10px;
        padding: 12px 16px;
        border-left: 4px solid #2980b9;
    }
    .alert-box {
        background: #fff3cd;
        border: 2px solid #f39c12;
        border-radius: 10px;
        padding: 14px 18px;
        margin: 10px 0;
    }
    .alert-box h4 { color: #7a5200; margin: 0 0 6px 0; }
    .alert-box p  { color: #9a6800; margin: 0; font-size: 13px; }
</style>
""", unsafe_allow_html=True)

RATIO = 2.5  # アラート閾値（在庫数 ≤ アベレージ × 2.5）

# ── Excel 解析 ────────────────────────────────────────────────────────
def is_error(v):
    if v is None: return True
    s = str(v).strip()
    return s in ("","None","NaN") or s.startswith('#')

def clean_num(v):
    if is_error(v): return None
    try:
        return int(round(float(v)))
    except:
        return None

def normalize(v):
    if v is None: return ''
    return str(v).replace('\n', ' ').replace('\r', '').strip()

@st.cache_data(show_spinner=False)
def parse_excel(file_bytes: bytes, filename: str):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_name = '在庫予定' if '在庫予定' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, "シートにデータがありません"

    maxcol = max(len(r) for r in rows)
    headers = [normalize(rows[0][c]) if c < len(rows[0]) else '' for c in range(maxcol)]

    code_col = next((i for i, h in enumerate(headers) if h == '商品コード'), 2)
    name_col = next((i for i, h in enumerate(headers) if h == '商品名'), 3)
    avr_col  = next((i for i, h in enumerate(headers)
                     if ('Avr' in h or 'avr' in h) and '実績' not in h), -1)
    hoyu_col = next((i for i, h in enumerate(headers)
                     if '保有在庫日数' in h or '保有日数' in h), -1)
    stock_date_cols = [
        (i, h) for i, h in enumerate(headers)
        if ('在庫' in h or '在庫数' in h) and '/' in h
        and '実績' not in h and '変動' not in h and '保有' not in h
    ]
    latest = stock_date_cols[-1] if stock_date_cols else None

    records = []
    for row in rows[1:]:
        def gc(c):
            return row[c] if 0 <= c < len(row) else None
        code = gc(code_col)
        if is_error(code): continue
        cs = str(code).strip()
        if not cs or cs in ('*', '**'): continue
        name = gc(name_col)
        records.append({
            '商品コード': cs,
            '商品名':     '' if is_error(name) else str(name).strip(),
            '在庫数':     clean_num(gc(latest[0])) if latest else None,
            'アベレージ': clean_num(gc(avr_col))   if avr_col  >= 0 else None,
            '保有日数':   clean_num(gc(hoyu_col))   if hoyu_col >= 0 else None,
        })

    meta = {
        'sheetName':        sheet_name,
        'latestStockLabel': latest[1] if latest else '在庫数',
        'totalStockCols':   len(stock_date_cols),
        'avrFound':         avr_col  >= 0,
        'hoyuFound':        hoyu_col >= 0,
    }
    return records, meta

def to_csv(df: pd.DataFrame) -> bytes:
    return ('\ufeff' + df.to_csv(index=False)).encode('utf-8')

def to_excel(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='在庫データ')
    return buf.getvalue()

# ── UI ───────────────────────────────────────────────────────────────
st.title("📦 在庫管理データ 抽出アプリ")
st.caption("Excelファイル（.xlsx / .xlsm / .xls）をアップロードするだけで自動解析します")

uploaded = st.file_uploader(
    "Excelファイルをアップロード",
    type=["xlsx", "xlsm", "xls"],
    label_visibility="collapsed"
)

if uploaded is None:
    st.info("⬆️ 上のエリアにExcelファイルをドラッグ＆ドロップ、またはクリックして選択してください")
    with st.expander("🔍 自動検出する列"):
        st.markdown("""
| 列名 | 検出ルール |
|------|-----------|
| 商品コード | 「商品コード」ヘッダー列 |
| 商品名 | 「商品名」ヘッダー列 |
| 在庫数（最新） | 日付付き在庫列の最右（例: 3/11 在庫数） |
| アベレージ | 「Avr」を含む列（7期月Avr など） |
| 保有日数 | 「保有在庫日数」列 |
""")
    st.stop()

# ── 解析 ─────────────────────────────────────────────────────────────
with st.spinner("読み込み中..."):
    file_bytes = uploaded.read()
    records, meta = parse_excel(file_bytes, uploaded.name)

if records is None:
    st.error(f"❌ {meta}")
    st.stop()

df = pd.DataFrame(records)

# ── ヘッダー情報 ──────────────────────────────────────────────────────
c1, c2, c3, c4 = st.columns(4)
c1.metric("📄 ファイル", uploaded.name.replace('.xlsm','').replace('.xlsx',''))
c2.metric("🗂 シート", meta['sheetName'])
c3.metric("📅 最新在庫", meta['latestStockLabel'])
c4.metric("📊 商品数", f"{len(df)} 件")

st.divider()

# ── アラート（在庫数 ≤ アベレージ × 2.5） ────────────────────────────
alert_df = df[
    df['在庫数'].notna() &
    df['アベレージ'].notna() &
    (df['アベレージ'] > 0) &
    (df['在庫数'] <= df['アベレージ'] * RATIO)
].copy()

if len(alert_df) > 0:
    alert_df['基準数量(Avr×2.5)'] = (alert_df['アベレージ'] * RATIO).apply(lambda x: int(round(x)))
    alert_df['不足数']             = alert_df['基準数量(Avr×2.5)'] - alert_df['在庫数']
    alert_df['倍率']               = (alert_df['在庫数'] / alert_df['アベレージ']).round(2)
    alert_df_sorted = alert_df.sort_values('倍率')

    st.markdown(f"""
    <div class="alert-box">
        <h4>⚠️ 過少在庫アラート：{len(alert_df)} 件</h4>
        <p>在庫数がアベレージの {RATIO} 倍以下の商品があります。下の「過少在庫アラート」タブで確認してください。</p>
    </div>
    """, unsafe_allow_html=True)

# ── 検索・フィルタ ────────────────────────────────────────────────────
search = st.text_input("🔍 商品コード・商品名で絞り込み", placeholder="キーワードを入力...")
if search:
    mask = (
        df['商品コード'].str.lower().str.contains(search.lower(), na=False) |
        df['商品名'].str.contains(search, na=False)
    )
    df_view = df[mask].copy()
else:
    df_view = df.copy()

st.caption(f"{len(df_view)} / {len(df)} 件表示中")

# ── タブ表示 ──────────────────────────────────────────────────────────
tabs = ["📋 在庫一覧"]
if len(alert_df) > 0:
    tabs.append(f"⚠️ 過少在庫アラート ({len(alert_df)}件)")

tab_list = st.tabs(tabs)

# 在庫一覧タブ
with tab_list[0]:
    # 色付き表示のためにstyleを使用
    def highlight_rows(row):
        s = row.get('在庫数', None)
        a = row.get('アベレージ', None)
        styles = [''] * len(row)
        col_names = list(row.index)

        if s is not None and s == 0:
            styles = ['background-color: #fff0f0; color: #e74c3c; font-weight: bold'] * len(row)
        elif s is not None and s <= 30:
            styles = ['background-color: #fff8f0; color: #e67e22; font-weight: bold'] * len(row)
        elif s is not None and a is not None and a > 0 and s <= a * RATIO:
            styles = ['background-color: #fffde7'] * len(row)
        return styles

    styled = df_view.style.apply(highlight_rows, axis=1)

    # 数値列はintegerフォーマット
    for col in ['在庫数', 'アベレージ', '保有日数']:
        if col in df_view.columns:
            styled = styled.format({col: lambda x: '' if pd.isna(x) else f'{int(x):,}'})

    st.dataframe(styled, use_container_width=True, height=500, hide_index=True)

    st.caption("🔴 欠品（在庫0）　🟠 僅少（30以下）　🟡 過少在庫（在庫数 ≤ アベレージ×2.5）")

    # ダウンロード
    dl_df = df_view[['商品コード', '商品名', '在庫数', 'アベレージ', '保有日数']].copy()
    # NaNを空文字に
    for col in ['在庫数', 'アベレージ', '保有日数']:
        dl_df[col] = dl_df[col].apply(lambda x: '' if pd.isna(x) else int(x))

    label = meta['latestStockLabel'].replace('/', '').replace(' ', '').replace('\n', '')
    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            "⬇️ CSVダウンロード",
            data=to_csv(dl_df),
            file_name=f"在庫データ_{label}.csv",
            mime="text/csv",
            use_container_width=True
        )
    with col2:
        st.download_button(
            "⬇️ Excelダウンロード",
            data=to_excel(dl_df),
            file_name=f"在庫データ_{label}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

# アラートタブ
if len(alert_df) > 0:
    with tab_list[1]:
        alert_show = alert_df_sorted[['商品コード', '商品名', '在庫数', 'アベレージ', '基準数量(Avr×2.5)', '不足数', '倍率']].copy()

        def highlight_alert(row):
            ratio = row.get('倍率', 1)
            if ratio < 0.5:
                return ['background-color: #ffebee'] * len(row)
            elif ratio < 1.0:
                return ['background-color: #fff3e0'] * len(row)
            else:
                return ['background-color: #fffde7'] * len(row)

        alert_styled = alert_show.style.apply(highlight_alert, axis=1)
        for col in ['在庫数', 'アベレージ', '基準数量(Avr×2.5)', '不足数']:
            alert_styled = alert_styled.format({col: lambda x: '' if pd.isna(x) else f'{int(x):,}'})
        alert_styled = alert_styled.format({'倍率': '{:.2f}倍'})

        st.dataframe(alert_styled, use_container_width=True, height=500, hide_index=True)
        st.caption(f"🔴 倍率 < 0.5　🟠 倍率 0.5〜1.0　🟡 倍率 1.0〜{RATIO} ／ 倍率が小さいほど在庫不足リスクが高い")

        # アラートCSVダウンロード
        alert_dl = alert_show.copy()
        for col in ['在庫数', 'アベレージ', '基準数量(Avr×2.5)', '不足数']:
            alert_dl[col] = alert_dl[col].apply(lambda x: '' if pd.isna(x) else int(x))
        alert_dl['倍率'] = alert_dl['倍率'].apply(lambda x: f'{x:.2f}')

        st.download_button(
            "⬇️ 過少在庫アラート CSV",
            data=to_csv(alert_dl),
            file_name=f"過少在庫アラート_{label}.csv",
            mime="text/csv",
            use_container_width=True
        )
